import tkinter as tk
from tkinter import filedialog, messagebox
import os
import shutil
import subprocess
import time
from datetime import timedelta
from openpyxl import load_workbook
import pytesseract
from PIL import Image
import platform

# Thiết lập đường dẫn Tesseract theo hệ điều hành
def detect_tesseract_path():
    system = platform.system()
    if system == "Windows":
        return r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    elif system == "Darwin":  # macOS
        # Ưu tiên M1/M2 Homebrew path
        if os.path.exists("/opt/homebrew/bin/tesseract"):
            return "/opt/homebrew/bin/tesseract"
        elif os.path.exists("/usr/local/bin/tesseract"):  # mac Intel
            return "/usr/local/bin/tesseract"
        else:
            return "tesseract"  # fallback, nếu đã có trong PATH
    else:
        return "tesseract"  # fallback cho Linux hoặc các HĐH khác

pytesseract.pytesseract.tesseract_cmd = detect_tesseract_path()

class ImageFilterApp:
    def __init__(self, master):
        self.master = master
        master.title("Lọc ảnh theo mã số by Chính HV")

        self.source_dir = ""
        self.target_dir = ""
        self.code_list_from_file = []

        self.main_frame = tk.Frame(master, padx=15, pady=15, bg="#2e2e2e")
        self.main_frame.pack(fill="both", expand=True)

        # Chọn thư mục
        self.frame_dirs = tk.LabelFrame(self.main_frame, text=" Chọn thư mục ", fg="white", bg="#2e2e2e", padx=10, pady=10)
        self.frame_dirs.pack(fill="x", pady=5)

        self.row1 = tk.Frame(self.frame_dirs, bg="#2e2e2e")
        self.row1.pack(fill="x", pady=2)
        tk.Label(self.row1, text="Thư mục ảnh gốc:", bg="#2e2e2e", fg="white", width=25, anchor="w").pack(side="left")
        self.label_source = tk.Label(self.row1, text="(Chưa chọn)", bg="#2e2e2e", fg="lightblue", anchor="w")
        self.label_source.pack(side="left", fill="x", expand=True, padx=5)
        tk.Button(self.row1, text="Duyệt", command=self.select_source).pack(side="right")

        self.row2 = tk.Frame(self.frame_dirs, bg="#2e2e2e")
        self.row2.pack(fill="x", pady=2)
        tk.Label(self.row2, text="Thư mục lưu ảnh đã lọc:", bg="#2e2e2e", fg="white", width=25, anchor="w").pack(side="left")
        self.label_target = tk.Label(self.row2, text="(Chưa chọn)", bg="#2e2e2e", fg="lightblue", anchor="w")
        self.label_target.pack(side="left", fill="x", expand=True, padx=5)
        self.button_target = tk.Button(self.row2, text="Duyệt", command=self.select_target)
        self.button_target.pack(side="right")

        # Checkbox for saving to source directory
        self.save_to_source_var = tk.BooleanVar(value=False)
        self.save_to_source_check = tk.Checkbutton(
            self.frame_dirs, text="Lưu vào thư mục gốc (File_Loc)", variable=self.save_to_source_var,
            command=self.toggle_target_selection, bg="#2e2e2e", fg="white", selectcolor="#444"
        )
        self.save_to_source_check.pack(anchor="w", pady=2, padx=10)

        # Nhập mã
        self.frame_codes = tk.LabelFrame(self.main_frame, text=" Nhập mã số hoặc chọn file ", fg="white", bg="#2e2e2e", padx=10, pady=10)
        self.frame_codes.pack(fill="x", pady=5)

        tk.Label(self.frame_codes, text="Nhập mã số (phân cách bởi dấu phẩy):", bg="#2e2e2e", fg="white").pack(anchor="w", padx=10)
        self.entry_code = tk.Text(self.frame_codes, width=80, height=5)
        self.entry_code.pack(pady=5, padx=10, fill="x")

        file_frame = tk.Frame(self.frame_codes, bg="#2e2e2e")
        file_frame.pack(fill="x", pady=5)
        tk.Button(file_frame, text="Chọn file mã (Excel hoặc TXT)", command=self.load_code_file).pack(side="left", padx=10)
        self.label_info = tk.Label(file_frame, text="Chưa chọn file mã số.", bg="#2e2e2e", fg="lightgreen")
        self.label_info.pack(side="left", padx=10)

        # OCR nhiều ảnh
        ocr_frame = tk.Frame(self.frame_codes, bg="#2e2e2e")
        ocr_frame.pack(fill="x", pady=5)
        tk.Button(ocr_frame, text="OCR nhiều ảnh", command=self.ocr_multiple_images).pack(side="left", padx=10)
        self.label_ocr_info = tk.Label(ocr_frame, text="", bg="#2e2e2e", fg="lightgreen")
        self.label_ocr_info.pack(side="left", padx=10)

        # Chọn định dạng lọc
        self.frame_format = tk.LabelFrame(self.main_frame, text=" Chọn định dạng cần lọc ", fg="white", bg="#2e2e2e", padx=10, pady=10)
        self.frame_format.pack(fill="x", pady=5)

        self.format_var = tk.StringVar(value="both")
        tk.Radiobutton(self.frame_format, text="Chỉ JPG", variable=self.format_var, value="jpg", bg="#2e2e2e", fg="white", selectcolor="#444").pack(side="left", padx=10)
        tk.Radiobutton(self.frame_format, text="Chỉ RAW", variable=self.format_var, value="raw", bg="#2e2e2e", fg="white", selectcolor="#444").pack(side="left", padx=10)
        tk.Radiobutton(self.frame_format, text="Cả hai", variable=self.format_var, value="both", bg="#2e2e2e", fg="white", selectcolor="#444").pack(side="left", padx=10)

        tk.Button(self.main_frame, text="Lọc Ảnh", command=self.filter_images, bg="green", fg="white").pack(fill="x", pady=10)

        # Kết quả
        self.result_label = tk.Label(self.main_frame, text="", bg="#2e2e2e", fg="white")
        self.result_label.pack(pady=5)

        # Tiến trình
        self.frame_progress = tk.LabelFrame(self.main_frame, text=" Tiến trình ", fg="white", bg="#2e2e2e")
        self.frame_progress.pack(fill="x", pady=5)

        self.label_time = tk.Label(self.frame_progress, text="🕒 Thời gian đã chạy: 0 giây", fg="lightgreen", bg="#2e2e2e")
        self.label_time.pack(anchor="w", padx=10)

        self.label_eta = tk.Label(self.frame_progress, text="⏳ Ước tính còn: đang tính toán...", fg="lightgreen", bg="#2e2e2e")
        self.label_eta.pack(anchor="w", padx=10)

        self.label_current = tk.Label(self.frame_progress, text="🖼️ Ảnh đang xử lý: (Chưa có)", fg="lightgreen", bg="#2e2e2e")
        self.label_current.pack(anchor="w", padx=10)

        self.label_done = tk.Label(self.frame_progress, text="✅ Đã xử lý: 0 / 0 mã", fg="lightgreen", bg="#2e2e2e")
        self.label_done.pack(anchor="w", padx=10)

        # Biến OCR
        self.ocr_codes = []

    def toggle_target_selection(self):
        if self.save_to_source_var.get():
            self.row2.pack_forget()  # Hide target directory selection
            if self.source_dir:
                self.target_dir = os.path.join(self.source_dir, "File_Loc")
                self.label_target.config(text=self.target_dir)
            else:
                self.label_target.config(text="(Chưa chọn thư mục gốc)")
        else:
            self.row2.pack(fill="x", pady=2, before=self.save_to_source_check)  # Show target directory selection above checkbox
            self.target_dir = ""
            self.label_target.config(text="(Chưa chọn)")

    def select_source(self):
        self.source_dir = filedialog.askdirectory()
        self.label_source.config(text=self.source_dir or "(Chưa chọn)")
        if self.save_to_source_var.get() and self.source_dir:
            self.target_dir = os.path.join(self.source_dir, "File_Loc")
            self.label_target.config(text=self.target_dir)
        elif self.save_to_source_var.get():
            self.label_target.config(text="(Chưa chọn thư mục gốc)")

    def select_target(self):
        if not self.save_to_source_var.get():
            self.target_dir = filedialog.askdirectory()
            self.label_target.config(text=self.target_dir or "(Chưa chọn)")

    def load_code_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("Text files", "*.txt")])
        if not file_path:
            return

        codes = []
        try:
            if file_path.endswith('.txt'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    codes = [line.strip() for line in f if line.strip()]
            elif file_path.endswith('.xlsx'):
                wb = load_workbook(filename=file_path)
                sheet = wb.active
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell:
                            codes.append(str(cell).strip())
        except Exception as e:
            messagebox.showerror("Lỗi đọc file", str(e))
            return

        self.code_list_from_file = list(set(codes))
        self.label_info.config(text=f"Đã nạp {len(self.code_list_from_file)} mã từ file.")

    def get_code_list(self):
        manual_input = self.entry_code.get("1.0", "end").strip()
        manual_codes = [c.strip() for c in manual_input.split(",") if c.strip()]
        if self.code_list_from_file:
            combined = set(self.code_list_from_file) | set(manual_codes)
            return list(combined)
        else:
            return manual_codes

    def normalize_string(self, s):
        return s.replace("-", "_").replace(" ", "").lower()

    def generate_unique_path(self, filename):
        dest_path = os.path.join(self.target_dir, filename)
        base, ext = os.path.splitext(filename)
        counter = 1
        while os.path.exists(dest_path):
            dest_path = os.path.join(self.target_dir, f"{base}_{counter}{ext}")
            counter += 1
        return dest_path

    def filter_images(self):
        if not self.source_dir:
            messagebox.showwarning("Thiếu thông tin", "Hãy chọn thư mục ảnh gốc.")
            return
        if not self.target_dir:
            messagebox.showwarning("Thiếu thông tin", "Hãy chọn thư mục lưu hoặc tích chọn 'Lưu vào thư mục gốc'.")
            return

        # Create File_Loc directory if save_to_source is checked
        if self.save_to_source_var.get():
            try:
                os.makedirs(self.target_dir, exist_ok=True)
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể tạo thư mục File_Loc: {str(e)}")
                return

        code_list = self.get_code_list()
        if not code_list:
            messagebox.showwarning("Thiếu mã số", "Hãy nhập mã hoặc chọn file.")
            return

        raw_exts = ['.cr2', '.cr3', '.nef', '.nrw', '.arw', '.srf', '.sr2',
                    '.raf', '.rw2', '.orf', '.pef', '.x3f', '.dng', '.3fr', '.iiq']
        jpg_exts = ['.jpg', '.jpeg']

        format_choice = self.format_var.get()
        extensions = {
            'jpg': jpg_exts,
            'raw': raw_exts,
            'both': jpg_exts + raw_exts
        }[format_choice]

        source_files = []
        for root, _, files in os.walk(self.source_dir):
            for file in files:
                full_path = os.path.join(root, file)
                source_files.append(full_path)

        start_time = time.time()
        matched_codes = 0
        copied_files = 0
        not_found = []
        failed_files = []

        for i, code in enumerate(code_list, 1):
            self.update_progress_ui(i, len(code_list), start_time, matched_codes)

            found = False
            for file_path in source_files:
                filename = os.path.basename(file_path)
                name, ext = os.path.splitext(filename)

                if (self.normalize_string(code) in self.normalize_string(name) and
                        ext.lower() in extensions):

                    try:
                        dest_path = self.generate_unique_path(filename)
                        shutil.copy2(file_path, dest_path)
                        copied_files += 1
                        found = True
                        break
                    except Exception as e:
                        failed_files.append(filename)
                        break

            if found:
                matched_codes += 1
            else:
                not_found.append(code)

        self.show_results(start_time, matched_codes, len(code_list), copied_files, not_found, failed_files)

        if os.path.exists(self.target_dir):
            try:
                if platform.system() == "Windows":
                    subprocess.Popen(["explorer", os.path.realpath(self.target_dir)])
                elif platform.system() == "Darwin":  # macOS
                    subprocess.Popen(["open", os.path.realpath(self.target_dir)])
            except Exception as e:
                messagebox.showwarning("Lỗi", f"Không thể mở thư mục lưu ảnh:\n{str(e)}")

    def update_progress_ui(self, current, total, start_time, matched):
        elapsed = time.time() - start_time
        eta = (elapsed / current) * (total - current) if current else 0
        self.label_time.config(text=f"🕒 Thời gian đã chạy: {int(elapsed)} giây")
        self.label_eta.config(text=f"⏳ Ước tính còn: {int(eta)} giây")
        self.label_done.config(text=f"✅ Đã xử lý: {matched} / {total} mã")
        self.label_current.config(text=f"🖼️ Ảnh đang xử lý: {current} / {total}")

    def show_results(self, start_time, matched_codes, total_codes, copied_files, not_found, failed_files):
        elapsed = timedelta(seconds=int(time.time() - start_time))
        result = (f"⏰ Thời gian chạy: {elapsed}\n"
                  f"✅ Mã tìm thấy: {matched_codes} / {total_codes}\n"
                  f"📂 Ảnh sao chép: {copied_files}\n"
                  f"⚠️ Mã không tìm thấy: {', '.join(not_found) if not_found else 'Không có'}\n"
                  f"❌ Ảnh lỗi khi sao chép: {', '.join(failed_files) if failed_files else 'Không có'}")
        self.result_label.config(text=result)

    def ocr_multiple_images(self):
        file_paths = filedialog.askopenfilenames(title="Chọn nhiều ảnh để OCR",
                                                 filetypes=[("Ảnh", "*.jpg *.jpeg *.png *.tif *.bmp *.arw *.nef *.cr2")])
        if not file_paths:
            return

        self.ocr_codes.clear()
        all_codes = []

        for idx, path in enumerate(file_paths, 1):
            self.label_ocr_info.config(text=f"Đang OCR ảnh {idx}/{len(file_paths)}: {os.path.basename(path)}")
            self.master.update()

            try:
                text = pytesseract.image_to_string(Image.open(path))
                codes = self.extract_codes_from_text(text)
                all_codes.extend(codes)
            except Exception as e:
                messagebox.showwarning("Lỗi OCR", f"Lỗi khi OCR ảnh {os.path.basename(path)}: {e}")

        unique_codes = list(set(all_codes))
        current_text = self.entry_code.get("1.0", "end").strip()
        if current_text:
            current_codes = [c.strip() for c in current_text.split(",") if c.strip()]
            combined = set(current_codes) | set(unique_codes)
        else:
            combined = set(unique_codes)

        self.entry_code.delete("1.0", "end")
        self.entry_code.insert("1.0", ", ".join(sorted(combined)))

        self.label_ocr_info.config(text=f"OCR hoàn thành, tìm được {len(unique_codes)} mã mới.")

    def extract_codes_from_text(self, text):
        import re
        codes = []

        # Normalize text: remove extra spaces, convert to lowercase
        text = ' '.join(text.split()).lower()

        # Pattern 1: Extract numbers from filenames like DSC8877.JPG or DSC_8877.jpg
        filename_pattern = r'(?:dsc_?)?(\d{3,10})(?=\.\w{3,4}\b)'
        filename_matches = re.findall(filename_pattern, text)
        codes.extend(filename_matches)

        # Pattern 2: Extract standalone number sequences (3-10 digits)
        number_pattern = r'\b(?<![\w-])([0-9]{3,10})(?![\w-])\b'
        number_matches = re.findall(number_pattern, text)
        
        # Filter out duplicates and numbers already found in filenames
        for num in number_matches:
            if num not in codes:
                codes.append(num)

        # Pattern 3: Handle custom code formats (e.g., ABC1234 or 1234-ABC)
        custom_pattern = r'\b[a-z0-9-]{3,10}\b'
        custom_matches = re.findall(custom_pattern, text)
        for match in custom_matches:
            if any(c.isdigit() for c in match) and not match.isalpha():
                if match not in codes:
                    codes.append(match)

        # Filter out false positives (e.g., years, short numbers)
        codes = [code for code in codes if not (
            re.match(r'^19\d{2}$|^20\d{2}$', code) or
            len(code) < 3
        )]

        return codes


if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("900x700")
    app = ImageFilterApp(root)
    root.mainloop()